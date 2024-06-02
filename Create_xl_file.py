
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Create file & sheets
wb = openpyxl.Workbook()
sheet1 = wb.active
sheet1.title = "Products"
wb.create_sheet("Cryovials")
sheet2 = wb["Cryovials"]

# Add column titles
sheet1.append(["Product", "Supplier", "Cat Number", "Expiration Date", "Quantity"])
sheet2.append(["Cell line", "Passage", "Date of freezing", "Frozen by", "Quantity"])
col = ["A", "B", "C", "D", "E"]

# Change title row style
for sheet in [sheet1, sheet2]:
    for x in col:
        sheet[f"{x}1"].font = Font(bold = True, size = 12, underline = "single")
        sheet[f"{x}1"].fill = PatternFill(fill_type= "solid", start_color= "00FFFFCC", end_color= "00FFFFCC")
        sheet[f"{x}1"].alignment = Alignment(horizontal= "center", vertical= "center")
        sheet.column_dimensions[x].width = 15
        

# Save file
wb.save("Lab_Inventory.xlsx")
print("Lab Inventory file created.")
